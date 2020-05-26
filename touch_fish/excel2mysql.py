#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time    : 2019/09/18 11:34
# @Author  : Shawn Wang

import os
import sys
import re
import pymysql
import openpyxl as xl


class Tool(object):
    def __init__(self, **kwargs):
        self.kw = kwargs

    def execute(self, sql, need_fetch=True, params=None):
        conn = None
        cur = None
        try:
            conn = pymysql.connect(**self.kw)
            cur = conn.cursor(pymysql.cursors.DictCursor)
            if isinstance(sql, str) and params is None:
                cur.execute(sql)
            elif isinstance(sql, str) and isinstance(params, (list, tuple)):
                cur.executemany(sql, params)
            elif isinstance(sql, list) and params is None:
                for s in sql:
                    cur.execute(s)

            if need_fetch:
                return cur.fetchall()
            else:
                conn.commit()
        except Exception as e:
            print(e.args[-1])
            conn.rollback()
            return ()
        finally:
            if cur:
                cur.close()
            if conn:
                conn.close()
        return True

    def obtain_file_path(self):
        for i in range(3):
            path = input('>Input: absolute file_path of .xlsx or .xls :')
            if os.path.exists(path):
                if path.endswith('.xlsx') or path.endswith('.xls'):
                    return path
                else:
                    print('>Error: Incorrect file format')
            else:
                print('>Error: File doesn\'t exist')
        else:
            print('>Fatal: More than three times,bye.')
            sys.exit(0)

    def obtain_table_field(self, tb_name):
        sql = """SHOW CREATE TABLE {}""".format(tb_name)
        res_sql = self.execute(sql)
        if res_sql:
            s = res_sql[0].get('Create Table')
            s_li = s.split('\n')
            s_li = filter(lambda x: x.startswith('`'), map(lambda x: x.lstrip(), s_li))
            return [[re.findall('[a-zA-Z]+', s)[0], re.findall('[a-zA-Z]+', s)[1]] for s in s_li]

    def obtain_xlsx_first_row(self, sheet, start=(1, 1)):
        row, col = start[1], start[0]
        table_head = []
        while True:
            data = sheet.cell(row=row, column=col).value
            if data:
                table_head.append([chr(col+64)+str(row), (row, col)])
            else:
                break
            col += 1
        return table_head

    def split_cells(self, sheet):
        for cell in sheet.merged_cell_ranges:
            tup = cell.bounds  # 假定A4:C5是合并的单元格，则得到(1,4,3,5)
            # s = cell.coord  # 得到的是一个字符串，如"A4:C5"
            row, col = tup[1], tup[0]
            data = sheet.cell(row=row, column=col).value  # 存储已合并单元格的原始数据
            sheet.unmerge_cells(cell.coord)  # 拆分单元格
            while col <= tup[2]:
                while row <= tup[3]:
                    sheet.cell(row=row, column=col).value = data
                    row += 1
                col += 1
        return sheet

    def confirm_column_name(self, table_field_list, xlsx_field_list, default=False):
        if len(table_field_list) == len(xlsx_field_list):
            if not default:
                temp = input('>Input: 是否需要指定列导入? 需要\'Y\',不需要Entry:')
                default = True if temp == '' else False
                if not default:
                    print('>Info: 数据库表的字段有：', end='')
                    for li in table_field_list:
                        print(li[0], end=',')
                    print('\b')
                    s = input('>Input: 请输入列名，并以逗号隔开(建议复制上面的):')
                    table_field = s.split(',')

                    print('>Info: xlsx表的字段有：', end='')
                    for li in xlsx_field_list:
                        print(li[0], end=',')
                    print('\b')
                    s = input('>Input: 请输入列名，并以逗号隔开(建议复制上面的):')
                    xlsx_field = s.split(',')
                else:
                    table_field = [li[0] for li in table_field_list]
                    xlsx_field = [li[0] for li in xlsx_field_list]
            else:
                table_field = [li[0] for li in table_field_list]
                xlsx_field = [li[0] for li in xlsx_field_list]
        else:
            print('>Info: 数据库表的字段有：', end='')
            for li in table_field_list:
                print(li[0], end=',')
            print('\b')
            s = input('>Input: 请输入列名，并以逗号隔开(建议复制上面的):')
            table_field = s.split(',')

            print('>Info: xlsx表的字段有：', end='')
            for li in xlsx_field_list:
                print(li[0], end=',')
            print('\b')
            s = input('>Input: 请输入列名，并以逗号隔开(建议复制上面的):')
            xlsx_field = s.split(',')

        return {
            'is_default': default,
            'table_field': table_field,
            'field_type': [li[1] for li in filter(lambda x: x[0] in table_field, table_field_list)],
            'xlsx_cell': xlsx_field,
            'xlsx_index': [(int(s[1:]), ord(s[:1])-64) for s in xlsx_field]
        }

    def track(self):
        path = self.obtain_file_path()
        wb = xl.load_workbook(path)
        sheet_list = wb.get_sheet_names()
        default = False
        for sheet_name in sheet_list:
            sheet = wb.get_sheet_by_name(sheet_name)
            tb_name = input('>Input: Loading [{}]..., 跳过请Entry,导入请指定table:'.format(sheet.title))
            if tb_name == '':
                print('>Info: Skip [{}]...'.format(sheet.title))
                continue
            else:
                # 开始对当前工作簿拆分合并的单元格(假如有的话)
                sheet = self.split_cells(sheet)
                confirm_dict = self.confirm_column_name(
                    self.obtain_table_field(tb_name),
                    self.obtain_xlsx_first_row(sheet),
                    default)
                # default = confirm_dict.get('is_default')
                # print(confirm_dict.get('table_field'))
                # print(confirm_dict.get('field_type'))
                # print(confirm_dict.get('xlsx_cell'))
                # print(confirm_dict.get('xlsx_index'))
                field = ['{},'.format(i) for i in confirm_dict.get('table_field')]
                field = ' '.join(field)
                field = field[:-1]

                cell_num = len(confirm_dict.get('xlsx_cell'))
                s = ['%s,' for i in range(cell_num)]
                s = ' '.join(s)
                s = s[:-1]

                row = 2
                data = []
                while 1:
                    v = sheet.cell(row=row, column=1).value
                    if not v:
                        break
                    data.append([sheet.cell(row=row, column=i).value for i in range(1, cell_num + 1)])
                    row += 1
                sql = """INSERT INTO {} ({}) VALUES ({})""".format(tb_name, field, s)
                print(sql)
                # self.execute(sql=sql, need_fetch=False, params=data)


if __name__ == '__main__':
    kw = {
        'host': '10.167.219.250',
        'user': 'npbg',
        'password': 'Fox1230!',
        'database': 'npbg',
        'charset': 'utf8mb4'
    }
    tool = Tool(**kw)
    tool.track()

